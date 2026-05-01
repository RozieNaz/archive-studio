import csv
import json
import re
import threading
import time
import textwrap
import urllib.parse
import urllib.request
from pathlib import Path
from tkinter import BOTH, BOTTOM, END, LEFT, RIGHT, VERTICAL, X, Y, BooleanVar, Button, Checkbutton, Entry, Frame, Label, Scrollbar, StringVar, Text, Tk, filedialog, messagebox
from tkinter.ttk import Progressbar, Style, Treeview

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


COLUMNS = [
    "Title",
    "Author",
    "DOI",
    "ISBN",
    "Suggested Filename",
    "Bibliography",
    "Confidence",
]

SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".epub",
    ".mobi",
    ".azw3",
    ".djvu",
    ".doc",
    ".docx",
    ".rtf",
    ".txt",
}

REQUEST_DELAY_SECONDS = 0.3
MAX_RESULTS = 5


def main():
    app = ArchiveStudioApp()
    app.mainloop()


class ArchiveStudioApp(Tk):
    def __init__(self):
        super().__init__()
        self.title("Archive Studio")
        self.geometry("1200x680")
        self.minsize(980, 520)
        self.configure(bg="#f6f3ee")

        self.folder_path = None
        self.rows = []
        self.item_to_row = {}
        self.status = StringVar(value="Choose a folder to begin.")
        self.wrap_text = BooleanVar(value=True)
        self.scan_progress = StringVar(value="")
        self.edit_vars = {}
        self.edit_texts = {}

        self._build_style()
        self._build_ui()

    def _build_style(self):
        self.style = Style(self)
        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        self.colors = {
            "bg": "#f6f3ee",
            "panel": "#fffdf8",
            "ink": "#202124",
            "muted": "#6f675c",
            "line": "#ddd5c9",
            "accent": "#9fdcc4",
            "accent_dark": "#75c8aa",
            "accent_text": "#12372d",
            "table_head": "#ece4d8",
            "select": "#d7ece8",
        }

        self.style.configure(
            "Archive.Treeview",
            background=self.colors["panel"],
            foreground=self.colors["ink"],
            fieldbackground=self.colors["panel"],
            bordercolor=self.colors["line"],
            rowheight=58,
            font=("Segoe UI", 9),
        )
        self.style.configure(
            "Archive.Treeview.Heading",
            background=self.colors["table_head"],
            foreground=self.colors["ink"],
            relief="flat",
            font=("Segoe UI Semibold", 9),
        )
        self.style.map(
            "Archive.Treeview",
            background=[("selected", self.colors["select"])],
            foreground=[("selected", self.colors["ink"])],
        )

    def _build_ui(self):
        header = Frame(self, bg=self.colors["bg"])
        header.pack(fill="x", padx=16, pady=(14, 8))

        title_block = Frame(header, bg=self.colors["bg"])
        title_block.pack(side=LEFT, fill="x", expand=True)

        Label(
            title_block,
            text="Archive Studio",
            bg=self.colors["bg"],
            fg=self.colors["ink"],
            font=("Segoe UI Semibold", 18),
        ).pack(anchor="w")
        Label(
            title_block,
            textvariable=self.status,
            bg=self.colors["bg"],
            fg=self.colors["muted"],
            font=("Segoe UI", 9),
            anchor="w",
        ).pack(anchor="w", pady=(2, 0))

        toolbar = Frame(self, bg=self.colors["bg"])
        toolbar.pack(fill="x", padx=16, pady=(0, 12))

        self._toolbar_button(toolbar, "Choose Folder & Scan", self.choose_folder).pack(side=LEFT, padx=(0, 8))
        self._toolbar_button(toolbar, "Fetch Metadata", self.fetch_metadata).pack(side=LEFT, padx=(0, 8))
        self._toolbar_button(toolbar, "Export XLSX", self.export_xlsx).pack(side=LEFT, padx=(0, 8))
        self._toolbar_button(toolbar, "Export CSV", self.export_csv).pack(side=LEFT, padx=(0, 8))

        Checkbutton(
            toolbar,
            text="Wrap Text",
            variable=self.wrap_text,
            command=self.refresh_table,
            bg=self.colors["bg"],
            fg=self.colors["ink"],
            activebackground=self.colors["bg"],
            selectcolor=self.colors["panel"],
            font=("Segoe UI", 9),
        ).pack(side=LEFT, padx=(8, 0))

        self.progress_frame = Frame(self, bg=self.colors["bg"])

        self.progress = Progressbar(self.progress_frame, orient="horizontal", mode="determinate", maximum=100)
        self.progress.pack(side=LEFT, fill="x", expand=True)

        Label(
            self.progress_frame,
            textvariable=self.scan_progress,
            bg=self.colors["bg"],
            fg=self.colors["muted"],
            font=("Segoe UI", 9),
            width=18,
            anchor="e",
        ).pack(side=RIGHT, padx=(10, 0))

        self.table_frame = Frame(self, bg=self.colors["panel"], highlightthickness=1, highlightbackground=self.colors["line"])
        self.table_frame.pack(fill=BOTH, expand=True, padx=16, pady=(0, 16))

        scrollbar = Scrollbar(self.table_frame, orient=VERTICAL)
        scrollbar.pack(side=RIGHT, fill=Y)

        xscrollbar = Scrollbar(self.table_frame, orient="horizontal")
        xscrollbar.pack(side=BOTTOM, fill=X)

        self.table = Treeview(
            self.table_frame,
            columns=COLUMNS,
            show="headings",
            yscrollcommand=scrollbar.set,
            xscrollcommand=xscrollbar.set,
            style="Archive.Treeview",
        )
        scrollbar.config(command=self.table.yview)
        xscrollbar.config(command=self.table.xview)

        widths = {
            "Title": 340,
            "Author": 240,
            "DOI": 190,
            "ISBN": 240,
            "Suggested Filename": 460,
            "Bibliography": 720,
            "Confidence": 120,
        }

        min_widths = {
            "Title": 52,
            "Author": 66,
            "DOI": 42,
            "ISBN": 46,
            "Suggested Filename": 142,
            "Bibliography": 100,
            "Confidence": 88,
        }

        for column in COLUMNS:
            self.table.heading(column, text=column)
            self.table.column(column, width=widths[column], minwidth=min_widths[column], stretch=False)

        self.table.pack(fill=BOTH, expand=True)
        self.table.bind("<Double-1>", self.show_selected_details)
        self.table.bind("<<TreeviewSelect>>", self.show_selected_details)
        self.table.bind("<ButtonRelease-1>", self.refresh_after_resize)

        detail_frame = Frame(self, bg=self.colors["panel"], highlightthickness=1, highlightbackground=self.colors["line"])
        detail_frame.pack(fill="x", padx=16, pady=(0, 16))

        Label(
            detail_frame,
            text="Selected Entry",
            bg=self.colors["panel"],
            fg=self.colors["muted"],
            font=("Segoe UI Semibold", 9),
            anchor="w",
        ).pack(fill="x", padx=10, pady=(8, 0))

        detail_buttons = Frame(detail_frame, bg=self.colors["panel"])
        detail_buttons.pack(fill="x", padx=10, pady=(6, 0))

        self._toolbar_button(detail_buttons, "Copy Suggested Filename", self.copy_suggested_filename).pack(side=LEFT, padx=(0, 8))
        self._toolbar_button(detail_buttons, "Copy Bibliography", self.copy_bibliography).pack(side=LEFT, padx=(0, 8))

        self.details = Text(
            detail_frame,
            height=6,
            wrap="word",
            bg=self.colors["panel"],
            fg=self.colors["ink"],
            relief="flat",
            font=("Segoe UI", 9),
        )
        self.details.pack(fill="x", padx=10, pady=(2, 8))
        self.details.insert("1.0", "Select any row to see the full entry here.")
        self.details.configure(state="disabled")

        edit_frame = Frame(detail_frame, bg=self.colors["panel"])
        edit_frame.pack(fill="x", padx=10, pady=(0, 10))

        for label in ["Title", "Author", "DOI", "ISBN", "Suggested Filename", "Confidence"]:
            row_frame = Frame(edit_frame, bg=self.colors["panel"])
            row_frame.pack(fill="x", pady=2)
            Label(row_frame, text=label, bg=self.colors["panel"], fg=self.colors["muted"], width=18, anchor="w").pack(side=LEFT)
            var = StringVar()
            self.edit_vars[label] = var
            Entry(row_frame, textvariable=var, bg="white", fg=self.colors["ink"], relief="solid", bd=1).pack(side=LEFT, fill="x", expand=True)

        bib_frame = Frame(edit_frame, bg=self.colors["panel"])
        bib_frame.pack(fill="x", pady=2)
        Label(bib_frame, text="Bibliography", bg=self.colors["panel"], fg=self.colors["muted"], width=18, anchor="nw").pack(side=LEFT)
        bibliography_box = Text(bib_frame, height=4, wrap="word", bg="white", fg=self.colors["ink"], relief="solid", bd=1)
        bibliography_box.pack(side=LEFT, fill="x", expand=True)
        self.edit_texts["Bibliography"] = bibliography_box

        self._toolbar_button(edit_frame, "Save Edits", self.save_selected_edits).pack(anchor="e", pady=(6, 0))

    def _toolbar_button(self, parent, text, command):
        return Button(
            parent,
            text=text,
            command=command,
            bg=self.colors["accent"],
            fg=self.colors["accent_text"],
            activebackground=self.colors["accent_dark"],
            activeforeground=self.colors["accent_text"],
            relief="flat",
            padx=12,
            pady=7,
            font=("Segoe UI Semibold", 9),
            cursor="hand2",
        )

    def show_progress(self):
        self.progress_frame.pack(fill="x", padx=16, pady=(0, 10), before=self.table_frame)

    def hide_progress(self):
        self.progress_frame.pack_forget()

    def choose_folder(self):
        selected = filedialog.askdirectory(title="Choose archive folder")
        if not selected:
            return
        self.folder_path = Path(selected)
        self.status.set(f"Selected: {self.folder_path}")
        self.scan_files()

    def scan_files(self):
        if not self.folder_path:
            self.choose_folder()
        if not self.folder_path:
            return

        thread = threading.Thread(target=self._scan_files_worker, daemon=True)
        thread.start()

    def _scan_files_worker(self):
        self.after(0, self.show_progress)
        self.after(0, self.status.set, "Scanning files...")
        self.after(0, self.scan_progress.set, "Preparing")
        self.after(0, self.progress.configure, {"value": 0})

        paths = [path for path in sorted(self.folder_path.rglob("*")) if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS]
        total = len(paths)
        existing = {row.get("Filename", "") for row in self.rows if row.get("Filename")}
        added = 0
        for index, path in enumerate(paths, start=1):
            cleaned = clean_filename_for_log(path.name)
            if cleaned in existing:
                percent = int((index / total) * 100) if total else 100
                self.after(0, self.progress.configure, {"value": percent})
                self.after(0, self.scan_progress.set, f"{index}/{total}")
                continue

            self.rows.append(
                {
                    "Filename": cleaned,
                    "Title": "",
                    "Author": "",
                    "DOI": "",
                    "ISBN": "",
                    "Suggested Filename": title_case_filename(cleaned),
                    "Bibliography": "",
                    "Confidence": "",
                }
            )
            existing.add(cleaned)
            added += 1
            percent = int((index / total) * 100) if total else 100
            self.after(0, self.progress.configure, {"value": percent})
            self.after(0, self.scan_progress.set, f"{index}/{total}")

        self.after(0, self.refresh_table)
        self.after(0, self.status.set, f"Scan complete. Added {added} new files. Total: {len(self.rows)}.")
        self.after(0, self.hide_progress)

    def fetch_metadata(self):
        if not self.rows:
            messagebox.showinfo("Archive Studio", "Scan files first.")
            return

        thread = threading.Thread(target=self._fetch_metadata_worker, daemon=True)
        thread.start()

    def _fetch_metadata_worker(self):
        for index, row in enumerate(self.rows):
            if row.get("Title") and row.get("Confidence"):
                continue

            query = clean_search_query(row["Filename"])
            candidates = []
            for search_query in search_query_variants(query):
                candidates.extend(search_google_books(search_query))
                time.sleep(REQUEST_DELAY_SECONDS)
                candidates.extend(search_open_library(search_query))
                time.sleep(REQUEST_DELAY_SECONDS)
                candidates.extend(search_crossref(search_query))
                time.sleep(REQUEST_DELAY_SECONDS)
                candidates.extend(search_crossref_title(search_query))
                time.sleep(REQUEST_DELAY_SECONDS)
                candidates.extend(search_openalex(search_query))
                time.sleep(REQUEST_DELAY_SECONDS)

            best = choose_best_candidate(row["Filename"], query, candidates)
            if best:
                row.update(
                    normalise_row_text(
                        {
                            "Title": best["title"],
                            "Author": best["author"],
                            "DOI": best["doi"],
                            "ISBN": best["isbn"],
                            "Suggested Filename": make_suggested_filename(best["author"], best["title"], best["year"]),
                            "Bibliography": make_bibliography(best),
                            "Confidence": best["confidence"],
                        }
                    )
                )
            else:
                row["Confidence"] = "Zero"

            self.after(0, self.refresh_table)
            self.after(0, self.status.set, f"Fetched metadata {index + 1}/{len(self.rows)}")

        self.after(0, self.status.set, "Metadata fetch complete.")

    def refresh_table(self):
        self.table.delete(*self.table.get_children())
        self.item_to_row = {}
        rowheight = 82 if self.wrap_text.get() else 34
        self.style.configure("Archive.Treeview", rowheight=rowheight)
        for index, row in enumerate(self.rows):
            item_id = self.table.insert("", END, values=[self.display_value(column, row.get(column, "")) for column in COLUMNS])
            self.item_to_row[item_id] = index

    def display_value(self, column, value):
        text = str(value or "")
        if not self.wrap_text.get():
            return text

        width_px = int(self.table.column(column, option="width") or 160)
        chars_per_line = max(10, int(width_px / 7.2))
        max_lines = 3
        return "\n".join(textwrap.wrap(text, width=chars_per_line, max_lines=max_lines, placeholder="..."))

    def refresh_after_resize(self, event=None):
        if self.wrap_text.get():
            self.after(50, self.refresh_table)

    def truncate(self, text, limit):
        if len(text) <= limit:
            return text
        return text[: max(0, limit - 3)].rstrip() + "..."

    def show_selected_details(self, event=None):
        selected = self.table.selection()
        if not selected:
            return
        row_index = self.item_to_row.get(selected[0])
        if row_index is None:
            return

        row = self.rows[row_index]
        lines = []
        for column in COLUMNS:
            value = row.get(column, "")
            if value:
                lines.append(f"{column}: {value}")

        self.details.configure(state="normal")
        self.details.delete("1.0", END)
        self.details.insert("1.0", "\n\n".join(lines) if lines else "No details available.")
        self.details.configure(state="disabled")
        self.load_selected_into_editor(row)

    def load_selected_into_editor(self, row):
        for column, var in self.edit_vars.items():
            var.set(row.get(column, ""))
        bibliography_box = self.edit_texts.get("Bibliography")
        if bibliography_box:
            bibliography_box.delete("1.0", END)
            bibliography_box.insert("1.0", row.get("Bibliography", ""))

    def save_selected_edits(self):
        row = self.selected_row()
        if not row:
            return
        for column, var in self.edit_vars.items():
            row[column] = var.get().strip()
        bibliography_box = self.edit_texts.get("Bibliography")
        if bibliography_box:
            row["Bibliography"] = bibliography_box.get("1.0", END).strip()
        row.update(normalise_row_text(row))
        self.refresh_table()
        self.status.set("Saved edits.")

    def selected_row(self):
        selected = self.table.selection()
        if not selected:
            return None
        row_index = self.item_to_row.get(selected[0])
        if row_index is None:
            return None
        return self.rows[row_index]

    def copy_suggested_filename(self):
        row = self.selected_row()
        if not row:
            return
        self.copy_to_clipboard(row.get("Suggested Filename", ""))

    def copy_bibliography(self):
        row = self.selected_row()
        if not row:
            return
        self.copy_to_clipboard(row.get("Bibliography", ""))

    def copy_to_clipboard(self, value):
        self.clipboard_clear()
        self.clipboard_append(str(value or ""))
        self.status.set("Copied to clipboard.")

    def export_xlsx(self):
        if Workbook is None:
            messagebox.showerror("Archive Studio", "openpyxl is not installed. Use CSV export instead.")
            return
        if not self.rows:
            messagebox.showinfo("Archive Studio", "Nothing to export.")
            return

        path = filedialog.asksaveasfilename(
            title="Save spreadsheet",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="metadata-log.xlsx",
        )
        if not path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metadata Log"
        sheet.append(COLUMNS)
        for row in self.rows:
            sheet.append([row.get(column, "") for column in COLUMNS])
        self.format_workbook_(sheet)
        workbook.save(path)
        self.status.set(f"Saved: {path}")

    def format_workbook_(self, sheet):
        from openpyxl.styles import Alignment, Font, PatternFill

        widths = {
            "A": 36,
            "B": 26,
            "C": 24,
            "D": 24,
            "E": 38,
            "F": 72,
            "G": 14,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        header_fill = PatternFill("solid", fgColor="ECE4D8")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def export_csv(self):
        if not self.rows:
            messagebox.showinfo("Archive Studio", "Nothing to export.")
            return

        path = filedialog.asksaveasfilename(
            title="Save CSV",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="metadata-log.csv",
        )
        if not path:
            return

        with open(path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=COLUMNS)
            writer.writeheader()
            for row in self.rows:
                writer.writerow({column: row.get(column, "") for column in COLUMNS})
        self.status.set(f"Saved: {path}")


def search_google_books(query):
    url = "https://www.googleapis.com/books/v1/volumes?q={}&maxResults={}".format(
        urllib.parse.quote(query), MAX_RESULTS
    )
    data = fetch_json(url)
    if not data or not data.get("items"):
        return []

    results = []
    for item in data["items"]:
        info = item.get("volumeInfo", {})
        identifiers = info.get("industryIdentifiers", [])
        isbn = ", ".join(identifier.get("identifier", "") for identifier in identifiers if identifier.get("identifier"))
        results.append(
            {
                "source": "Google Books",
                "type": "book",
                "title": clean_text(info.get("title", "")),
                "author": ", ".join(info.get("authors", [])),
                "year": extract_year(info.get("publishedDate", "")),
                "publisher": clean_text(info.get("publisher", "")),
                "publication": "",
                "doi": "",
                "isbn": isbn,
            }
        )
    return [item for item in results if item["title"]]


def search_open_library(query):
    url = "https://openlibrary.org/search.json?q={}&limit={}".format(urllib.parse.quote(query), MAX_RESULTS)
    data = fetch_json(url)
    if not data or not data.get("docs"):
        return []

    results = []
    for item in data["docs"]:
        results.append(
            {
                "source": "Open Library",
                "type": "book",
                "title": clean_text(item.get("title", "")),
                "author": ", ".join(item.get("author_name", [])[:3]),
                "year": item.get("first_publish_year", ""),
                "publisher": clean_text((item.get("publisher") or [""])[0]),
                "publication": "",
                "doi": "",
                "isbn": ", ".join((item.get("isbn") or [])[:3]),
            }
        )
    return [item for item in results if item["title"]]


def search_crossref(query):
    url = "https://api.crossref.org/works?query.bibliographic={}&rows={}".format(
        urllib.parse.quote(query), MAX_RESULTS
    )
    data = fetch_json(url)
    if not data or not data.get("message", {}).get("items"):
        return []

    results = []
    for item in data["message"]["items"]:
        results.append(
            {
                "source": "Crossref",
                "type": item.get("type", ""),
                "title": clean_text((item.get("title") or [""])[0]),
                "author": format_crossref_authors(item.get("author", [])),
                "year": extract_crossref_year(item),
                "publisher": clean_text(item.get("publisher", "")),
                "publication": clean_text((item.get("container-title") or [""])[0]),
                "doi": item.get("DOI", ""),
                "isbn": ", ".join(item.get("ISBN", [])),
            }
        )
    return [item for item in results if item["title"]]


def search_crossref_title(query):
    url = "https://api.crossref.org/works?query.title={}&rows={}".format(
        urllib.parse.quote(query), MAX_RESULTS
    )
    data = fetch_json(url)
    if not data or not data.get("message", {}).get("items"):
        return []

    results = []
    for item in data["message"]["items"]:
        results.append(
            {
                "source": "Crossref Title",
                "type": item.get("type", ""),
                "title": clean_text((item.get("title") or [""])[0]),
                "author": format_crossref_authors(item.get("author", [])),
                "year": extract_crossref_year(item),
                "publisher": clean_text(item.get("publisher", "")),
                "publication": clean_text((item.get("container-title") or [""])[0]),
                "doi": item.get("DOI", ""),
                "isbn": ", ".join(item.get("ISBN", [])),
            }
        )
    return [item for item in results if item["title"]]


def search_openalex(query):
    url = "https://api.openalex.org/works?search={}&per-page={}".format(urllib.parse.quote(query), MAX_RESULTS)
    data = fetch_json(url)
    if not data or not data.get("results"):
        return []

    results = []
    for item in data["results"]:
        authors = []
        for authorship in item.get("authorships", []):
            author = authorship.get("author", {})
            if author.get("display_name"):
                authors.append(author["display_name"])
        source = ((item.get("primary_location") or {}).get("source") or {}).get("display_name", "")
        doi = (item.get("doi") or "").replace("https://doi.org/", "")
        results.append(
            {
                "source": "OpenAlex",
                "type": item.get("type", ""),
                "title": clean_text(item.get("display_name", "")),
                "author": ", ".join(authors[:3]),
                "year": item.get("publication_year", ""),
                "publisher": "",
                "publication": clean_text(source),
                "doi": doi,
                "isbn": "",
            }
        )
    return [item for item in results if item["title"]]


def choose_best_candidate(filename, query, candidates):
    if not candidates:
        return None
    for candidate in candidates:
        score, confidence = score_candidate(filename, query, candidate)
        candidate["score"] = score
        candidate["confidence"] = confidence
    candidates.sort(key=lambda item: item["score"], reverse=True)
    if candidates[0]["confidence"] == "Zero":
        candidates[0]["confidence"] = "Low"
    return enrich_candidate(candidates[0], candidates)


def enrich_candidate(best, candidates):
    for candidate in candidates:
        if not is_same_work(best, candidate):
            continue
        if not best.get("author") and candidate.get("author"):
            best["author"] = candidate["author"]
        if not best.get("isbn") and candidate.get("isbn"):
            best["isbn"] = candidate["isbn"]
        if not best.get("doi") and candidate.get("doi"):
            best["doi"] = candidate["doi"]
        if not best.get("publisher") and candidate.get("publisher"):
            best["publisher"] = candidate["publisher"]
        if not best.get("publication") and candidate.get("publication"):
            best["publication"] = candidate["publication"]
        if not best.get("year") and candidate.get("year"):
            best["year"] = candidate["year"]
    return best


def is_same_work(left, right):
    if left is right:
        return True
    if left.get("doi") and right.get("doi") and left["doi"].lower() == right["doi"].lower():
        return True
    if left.get("isbn") and right.get("isbn") and shared_identifier(left["isbn"], right["isbn"]):
        return True
    return token_overlap(normalise_for_match(left.get("title", "")), normalise_for_match(right.get("title", ""))) >= 0.72


def shared_identifier(left, right):
    left_ids = {re.sub(r"[^0-9xX]", "", value) for value in str(left).split(",")}
    right_ids = {re.sub(r"[^0-9xX]", "", value) for value in str(right).split(",")}
    left_ids.discard("")
    right_ids.discard("")
    return bool(left_ids & right_ids)


def score_candidate(filename, query, candidate):
    file_text = normalise_for_match(query or filename)
    title_text = normalise_for_match(candidate["title"])
    author_text = normalise_for_match(candidate["author"])
    score = token_overlap(file_text, title_text) * 60

    if candidate["author"] and contains_any_token(file_text, author_text):
        score += 15
    if candidate["isbn"]:
        score += 8
    if candidate["doi"]:
        score += 5
    if candidate["year"] and str(candidate["year"]) in filename:
        score += 8
    if is_bad_title(candidate["title"]):
        score -= 20
    if is_likely_wrong_academic_result(candidate, filename):
        score -= 12

    if score >= 55:
        confidence = "High"
    elif score >= 32:
        confidence = "Medium"
    elif score >= 12:
        confidence = "Low"
    else:
        confidence = "Zero"

    return max(0, round(score)), confidence


def clean_filename_for_log(filename):
    text = Path(filename).stem
    text = re.sub(r"\b(z[-_\s]*library|zlib|1lib|libgen|pdf\s*drive|pdfdrive|anna'?s?\s*archive|archive\.org)\b", " ", text, flags=re.I)
    text = re.sub(r"\b(single|page|scan|scanned|ocr|retail|ebook|e-book|epub|pdf|mobi|azw3)\b", " ", text, flags=re.I)
    text = re.sub(r"\[[^\]]*\]", " ", text)
    text = re.sub(r"\([^)]*(z[-_\s]*library|zlib|1lib|libgen|pdf\s*drive|pdfdrive|anna'?s?\s*archive|epub|pdf|mobi|azw3|ocr|scan)[^)]*\)", " ", text, flags=re.I)
    text = re.sub(r"[_+$@#%^~=]+", " ", text)
    text = re.sub(r"[{}\[\]|\\/:;\"<>?*!]+", " ", text)
    text = re.sub(r"[-–—]+", " - ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return title_case_filename(text)


def clean_search_query(filename):
    text = clean_filename_for_log(filename)
    text = re.sub(r"\b(second|third|fourth|fifth|edition|edn|vol|volume)\b", " ", text, flags=re.I)
    return re.sub(r"\s+", " ", text).strip()


def search_query_variants(query):
    variants = []
    base = re.sub(r"\s+", " ", str(query or "")).strip()
    if base:
        variants.append(base)

    if " - " in base:
        parts = [part.strip() for part in base.split(" - ") if part.strip()]
        if len(parts) >= 2:
            variants.append(parts[-1])
            variants.append(" ".join(reversed(parts[:2])))

    cleaned = re.sub(r"\b(review of|review|book review)\b", " ", base, flags=re.I)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if cleaned and cleaned != base:
        variants.append(cleaned)

    without_parentheses = re.sub(r"\([^)]*\)", " ", base)
    without_parentheses = re.sub(r"\s+", " ", without_parentheses).strip()
    if without_parentheses and without_parentheses != base:
        variants.append(without_parentheses)

    no_author_initials = re.sub(r"\b[A-Z]\s*\.?\s+(?=[A-Z][a-z])", " ", base)
    no_author_initials = re.sub(r"\s+", " ", no_author_initials).strip()
    if no_author_initials and no_author_initials != base:
        variants.append(no_author_initials)

    unique = []
    seen = set()
    for variant in variants:
        key = variant.lower()
        if key not in seen:
            unique.append(variant)
            seen.add(key)
    return unique[:5]


def make_suggested_filename(author, title, year):
    main_title = (title or "").split(":")[0].strip()
    first_author = (author or "").split(",")[0].strip()
    if first_author and main_title and year:
        value = f"{first_author} - {main_title} ({year})"
    elif first_author and main_title:
        value = f"{first_author} - {main_title}"
    elif main_title and year:
        value = f"{main_title} ({year})"
    else:
        value = main_title
    value = re.sub(r"[\/\\:*?\"<>|]", "", value)
    return title_case_filename(value)


def make_bibliography(item):
    parts = []
    if item["author"]:
        parts.append(item["author"] + ".")
    if item["title"]:
        parts.append(item["title"] + ".")
    if item["publication"]:
        parts.append(item["publication"] + ".")
    if item["publisher"]:
        parts.append(item["publisher"] + ".")
    if item["year"]:
        parts.append(str(item["year"]) + ".")
    if item["doi"]:
        parts.append("DOI: " + item["doi"] + ".")
    elif item["isbn"]:
        parts.append("ISBN: " + item["isbn"] + ".")
    return re.sub(r"\s+", " ", " ".join(parts)).strip()


def normalise_row_text(row):
    output = dict(row)
    for field in ["Title", "Author", "Suggested Filename", "Bibliography", "Confidence"]:
        if output.get(field):
            output[field] = title_case_text(output[field])
    return output


def title_case_text(text):
    placeholders = {}

    def protect(match):
        key = f"__KEEP{len(placeholders)}__"
        placeholders[key] = match.group(0)
        return key

    value = str(text or "")
    value = re.sub(r"\b10\.\d{4,9}/[^\s.]+", protect, value, flags=re.I)
    value = re.sub(r"\b(?:97[89][-\s]?)?(?:\d[-\s]?){9,12}[\dXx]\b", protect, value)
    value = re.sub(r"\bDOI\b", protect, value, flags=re.I)
    value = re.sub(r"\bISBN\b", protect, value, flags=re.I)

    cased = title_case_filename(value)
    for key, original in placeholders.items():
        cased = re.sub(re.escape(key), original, cased, flags=re.I)
    return cased


def title_case_filename(text):
    small_words = {"a", "an", "and", "as", "at", "but", "by", "for", "from", "in", "into", "nor", "of", "on", "or", "the", "to", "with"}
    words = str(text or "").lower().split()
    output = []
    for index, word in enumerate(words):
        if word == "-":
            output.append(word)
        elif 0 < index < len(words) - 1 and word in small_words:
            output.append(word)
        else:
            output.append(word[:1].upper() + word[1:])
    return re.sub(r"\s+", " ", " ".join(output)).strip()


def fetch_json(url):
    request = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": "ArchiveStudio/0.1"})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception:
        return None


def clean_text(value):
    text = re.sub(r"<[^>]+>", "", str(value or ""))
    text = text.replace("&amp;", "&").replace("&quot;", '"').replace("&#39;", "'")
    return re.sub(r"\s+", " ", text).strip()


def normalise_for_match(value):
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\b(the|a|an|and|or|of|in|on|to|for|with|by|from|edition|volume|vol)\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def token_overlap(a, b):
    a_tokens = set(unique_tokens(a))
    b_tokens = set(unique_tokens(b))
    if not a_tokens or not b_tokens:
        return 0
    return len(a_tokens & b_tokens) / max(len(a_tokens), len(b_tokens))


def unique_tokens(value):
    return [token for token in str(value or "").split() if len(token) >= 3]


def contains_any_token(haystack, needles):
    haystack_tokens = set(unique_tokens(haystack))
    return any(token in haystack_tokens for token in unique_tokens(needles))


def is_bad_title(title):
    text = normalise_for_match(title)
    return not text or text in {"intro", "introduction", "contents", "notes", "index", "bibliography", "references", "world"}


def is_likely_wrong_academic_result(item, filename):
    item_type = str(item.get("type", "")).lower()
    overlap = token_overlap(normalise_for_match(filename), normalise_for_match(item.get("title", "")))
    if re.search(r"journal-article|book-chapter|proceedings-article|posted-content", item_type) and overlap < 0.55:
        return True
    return bool(re.search(r"\b(review of|book review|chapter|introduction|contents)\b", item.get("title", ""), re.I))


def extract_year(value):
    match = re.search(r"\b(15|16|17|18|19|20)\d{2}\b", str(value or ""))
    return match.group(0) if match else ""


def extract_crossref_year(item):
    date_parts = (item.get("issued") or {}).get("date-parts") or []
    if date_parts and date_parts[0]:
        return date_parts[0][0]
    return ""


def format_crossref_authors(authors):
    names = []
    for author in authors or []:
        name = author.get("name", "")
        if not name:
            name = " ".join(part for part in [author.get("given", ""), author.get("family", "")] if part)
        if name:
            names.append(name)
    return ", ".join(names[:3])


if __name__ == "__main__":
    main()
